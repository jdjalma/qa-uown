#!/usr/bin/env python3
"""Idempotent updater for docs/dev3-falta-testar.xlsx.

Reads _status_updates.json (list of {"row": N, "status": "OK", "note": "..."})
and writes column F (Task Status) and column G (Notes) for each row.
Re-runnable; only touches listed rows. Read row map first with --list.

Usage:
  python3 _set_status.py --list
  python3 _set_status.py            # applies _status_updates.json
"""
import json, sys, os
import openpyxl

XLSX = os.path.join(os.path.dirname(__file__), "..", "..", "dev3-falta-testar.xlsx")
UPD = os.path.join(os.path.dirname(__file__), "_status_updates.json")

def main():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active
    if "--list" in sys.argv:
        for r in range(1, ws.max_row + 1):
            portal = ws.cell(r, 2).value or ""
            task = ws.cell(r, 3).value or ""
            status = ws.cell(r, 6).value or ""
            if task or portal:
                print(f"{r}\t{portal}\t{task}\t[{status}]")
        return
    with open(UPD) as f:
        updates = json.load(f)
    applied = 0
    for u in updates:
        r = int(u["row"])
        task = ws.cell(r, 3).value or ""
        # safety: skip header/empty rows
        if not task:
            print(f"SKIP row {r}: empty task cell", file=sys.stderr)
            continue
        ws.cell(r, 6).value = u.get("status", "OK")
        if u.get("note"):
            ws.cell(r, 7).value = u["note"]
        applied += 1
        print(f"row {r} [{task[:55]}] -> {u.get('status','OK')}")
    wb.save(XLSX)
    print(f"\nApplied {applied} updates -> {os.path.normpath(XLSX)}")

if __name__ == "__main__":
    main()
